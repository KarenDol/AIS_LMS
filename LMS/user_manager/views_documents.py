from .models import Student
from .const import *
from .dogovor import *
from .decorators import role_required
from .helpers import get_student_or_redirect
from datetime import datetime
from django.contrib import messages
from django.shortcuts import render, redirect
from django.conf import settings
from django.contrib.auth.models import User
from django.forms.models import model_to_dict
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from .views_curator import *
import json, re, os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from io import BytesIO


@role_required(USER_TYPE_VNSV, USER_TYPE_CURATOR)
def export(request, grade):
    if grade == "Все классы":
        students = list(Student.objects.select_related("parent", "contract")
                        .filter(status="Акт")
                        .order_by('grade_num', 'grade_let', 'Last_Name', 'First_Name', 'Patronim')
                        .values('Last_Name', 'First_Name', 'Patronim', 'IIN', 'grade_num', 'grade_let',
                                'parent__Last_Name', 'parent__First_Name', 'parent__Patronim',
                                'parent__Phone', 'parent__Address', 'contract__monthly'))
    else:
        match = re.match(r"(\d+)([A-ZА-ЯӘҒҚҢӨҰҮҺІ])", grade)
        grade_num = match.group(1)  
        grade_let = match.group(2)
        students = list(Student.objects.select_related("parent", "contract")
                        .filter(status="Акт", grade_num = grade_num, grade_let = grade_let)
                        .order_by('Last_Name', 'First_Name', 'Patronim')
                        .values('Last_Name', 'First_Name', 'Patronim', 'IIN', 'grade_num', 'grade_let',
                                'parent__Last_Name', 'parent__First_Name', 'parent__Patronim',
                                'parent__Phone', 'parent__Address', 'contract__monthly'))
        
    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(students)

    df['ФИО Ученика'] = df['Last_Name'] + ' ' + df['First_Name'] + ' ' + df['Patronim']
    df['ФИО Родителя'] = df['parent__Last_Name'] + ' ' + df['parent__First_Name'] + ' ' + df['parent__Patronim']
    df['Класс'] = df['grade_num'].astype(str) + ' ' + df['grade_let'] #stringify grade_num

    # Rename columns
    df = df.rename(columns={
        'IIN': 'ИИН',
        'parent__Phone': 'Номер',
        'parent__Address': 'Адрес',
        'contract__monthly': 'Оплата',
    })

    #Reorder the columns
    if grade == "Все классы":
        df = df[['ФИО Ученика', 'ИИН', 'ФИО Родителя', 'Номер', 'Адрес', 'Оплата', 'Класс']]
    else:
        df = df[['ФИО Ученика', 'ИИН', 'ФИО Родителя', 'Номер', 'Адрес', 'Оплата']]

    # Add index column
    df.insert(0, '#', range(1, len(df) + 1))
    # Ensure the index reflects this column visually
    df = df.set_index('#')

    # Save the DataFrame to an Excel file
    doc_location = os.path.join(settings.STATIC_ROOT, "user_manager", 'students.xlsx')
    df.to_excel(doc_location, index=True, sheet_name=grade)

    # Load the Excel file using openpyxl
    workbook = load_workbook(doc_location)
    sheet = workbook.active

    # Define a fill style for the header
    header_fill = PatternFill(start_color="59007A", end_color="59007A", fill_type="solid")  # Yellow color
    # Define a font style for the header (white text)
    header_font = Font(color="FFFFFF", bold=True)  # White color, bold
    # Define a border style
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Apply the fill style to the header row
    for cell in sheet[1]:  # Header is the first row
        cell.fill = header_fill
        cell.font = header_font

    # Apply borders to all cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border  # Apply border to every cell

    # Adjust the width of specific columns
    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['B'].width = 40   
    sheet.column_dimensions['C'].width = 12  
    sheet.column_dimensions['D'].width = 40 
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 5    

    # Add info about the class
    sheet['F'+str(len(df)+4)] = grade + ' - ' + str(len(df)) + ' ' + 'учеников'
    sheet['F'+str(len(df)+4)].font = Font(bold=True)
    sheet['F'+str(len(df)+4)].alignment = Alignment(horizontal="right")

    # Save the styled Excel file
    workbook.save(doc_location)
    return FileResponse(open(doc_location, 'rb'))
    

@role_required(USER_TYPE_VNSV)
def join_doc(request, IIN):
    student = get_student_or_redirect(IIN)
    if isinstance(student, redirect.__class__):  # If redirection is returned
        return student
    
    if request.session.get('user_type') == 'ВнСв':
        fill_join(IIN)
        file_path = os.path.join(settings.STATIC_ROOT, 'user_manager', 'docs', 'join' + IIN + '.docx')

        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                file_data = BytesIO(f.read())

            # Delete the file
            try:
                os.remove(file_path)
            except OSError:
                raise Http404("Unable to delete the file")
            
            # Return the file response
            file_data.seek(0)  # Reset buffer pointer to the start
            return FileResponse(file_data)
        else:
            raise Http404("File not found")
    else:
        messages.error(request, "Только зам по ВСиРШ можем менять карточку студента")
        return redirect('home')

@role_required(USER_TYPE_VNSV)
def leave_doc(request, IIN):
    student = get_student_or_redirect(IIN)
    if isinstance(student, redirect.__class__):
        return student
    
    fill_leave(IIN)
    file_path = os.path.join(settings.STATIC_ROOT, 'user_manager', 'docs', 'leave' + IIN + '.docx')

    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            file_data = BytesIO(f.read())

        # Delete the file
        try:
            os.remove(file_path)
        except OSError:
            raise Http404("Unable to delete the file")
        
        # Return the file response
        file_data.seek(0)  # Reset buffer pointer to the start
        return FileResponse(file_data)
    else:
        raise Http404("File not found")


def spravka(request):
    return render(request, 'user_manager/spravka.html')

def get_spravka(request):
    IIN = request.session["IIN"]
    PIN_status = request.session.get("PIN_status")

    if request.user.is_authenticated or PIN_status:
        fill_spravka(IIN)
        file_path = os.path.join(settings.STATIC_ROOT, 'user_manager', 'docs', 'spravka' + IIN + '.docx')

        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                file_data = BytesIO(f.read())

            # Delete the file
            try:
                os.remove(file_path)
            except OSError:
                raise Http404("Unable to delete the file")

            # Return the file response
            file_data.seek(0)  # Reset buffer pointer to the start
            return FileResponse(file_data)
        else:
            raise Http404("File not found")
    else:
        raise Http404("User is unauthenticated or phone number is not verified")

@role_required(USER_TYPE_VNSV, USER_TYPE_CURATOR)
def sign_doc(request, IIN):
    student = get_student_or_redirect(IIN)
    if isinstance(student, HttpResponse):
        messages.error(request, "Ученика с таким ИИН нет в системе")
        return student
    
    if (student.contract == None):
        messages.error(request, "Договора нет в системе")
        return redirect('register_contract', IIN=IIN)
    
    contract = student.contract

    if request.method == "POST":
        uploaded_file = request.FILES.get('file')  # 'file' matches the JS key

        if uploaded_file:
            # Construct the path to save the uploaded file
            file_path = os.path.join(settings.STATIC_ROOT, 'user_manager', 'docs', f'{contract.numb}.pdf')

            # Ensure the target directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # ❌ Delete existing file if it exists
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except OSError as e:
                    return JsonResponse({'status': 'error', 'message': f'Ошибка удаления файла: {str(e)}'})
            
            # Save the uploaded file to the path
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            
            contract.status = True
            contract.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No file received'})
        
    context = {
        "IIN": IIN,
        "numb": contract.numb,
        "status": contract.status
    }

    return render(request, 'user_manager/sign_doc.html', context)

def fill_contract(request, IIN, numb):
    fill_doc(IIN, request.session['school'])
    file_path = os.path.join(settings.STATIC_ROOT, 'user_manager', 'docs', 'dogovor' + str(numb) + '.docx')

    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            file_data = BytesIO(f.read())

        # Delete the file
        try:
            os.remove(file_path)
        except OSError:
            raise Http404("Unable to delete the file")
        
        # Return the file response
        file_data.seek(0)  # Reset buffer pointer to the start
        return FileResponse(file_data, as_attachment=True, filename='Договор_для_подписания.docx')
    else:
        raise Http404("File not found")